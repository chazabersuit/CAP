from datetime import datetime
from django.shortcuts import render
from .forms import Formulario
from .models import Prueba
from openpyxl import load_workbook

# Create your views here.
def inicio (request):
   fecha= datetime.now()
   return render(request, 'index.html',{'fecha':fecha})

def crear_servicios(request):
   # nombre= request.POST.get('nombre')
   # edad= request.POST.get('edad')
   # datos = Prueba(nombre=nombre,edad=edad, fecha_creacion=datetime.now())
   # datos.save()
   miformulario1=Formulario()
   
   if request.method== 'POST':
      miformulario=Formulario(request.POST)
      
      if miformulario.is_valid():
         data= miformulario.cleaned_data
         datos= Prueba(nombre=data.get('nombre'),edad=data.get('edad'), fecha_creacion=datetime.now())
         datos.save()
         mensaje='Guardado con exito'
         datos= Prueba.objects.all()
         return render(request,'Servicios/listado_servicios.html', {'mensaje':mensaje,'datos':datos})
      else:
         return render(request,'Servicios/crear_servicios.html',{'miformulario':miformulario})
      
   return render(request,'Servicios/crear_servicios.html',{'miformulario':miformulario1})

def listado_servicios(request):
   datos= Prueba.objects.all()
   return render(request,'Servicios/listado_servicios.html',{'datos':datos})

def listado_excel(request):
   archivo_excel=r'C:\Users\CHAZA\Documents\Python\CAP\2023.xlsx'
   
   libro= load_workbook(archivo_excel)
   hoja=libro['Hoja1']
   datos = []
   
   for fila in hoja.iter_rows(values_only=True):
      datos.append(fila)
   
   titulo=datos.pop(0)
   return render(request,'Servicios/listado_excel.html',{'datos':datos,'titulo':titulo})